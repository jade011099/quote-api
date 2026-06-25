import os, json, io, requests
from datetime import date
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import load_workbook

app = Flask(__name__)
CORS(app)

GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")
GEMINI_MODEL = os.environ.get("GEMINI_MODEL", "gemini-3.1-pro-preview")
TEMPLATE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "報價單-空白單.xlsx")

OCR_PROMPT = """你是報價單辨識專家。請仔細辨識圖片中所有手寫或印刷文字，
輸出**純JSON**（不要markdown反引號、不要任何說明文字）：
{
  "client": "此致/客戶名稱",
  "phone": "客戶電話",
  "fax": "客戶傳真",
  "date": "報價日期YYYY-MM-DD（民國年請轉西元）",
  "site": "工地地點",
  "items": [
    {"name":"品名","spec":"規格","qty":數量,"unit":"單位","price":單價,"note":"備註"}
  ]
}
看不清楚的欄位填空字串或0。items至少一筆。"""


# ─── 路由 ─────────────────────────────────────────────────────────────

@app.route("/")
def health():
    return jsonify(status="ok", msg="報價單API運作中 ✅")


@app.route("/ocr", methods=["POST"])
def ocr():
    body = request.get_json(silent=True)
    if not body or "image" not in body:
        return jsonify(error="缺少 image 欄位"), 400
    if not GEMINI_API_KEY:
        return jsonify(error="伺服器未設定 GEMINI_API_KEY"), 500

    url = f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent"
    payload = {
        "contents": [{"parts": [
            {"text": OCR_PROMPT},
            {"inline_data": {
                "mime_type": body.get("mime", "image/jpeg"),
                "data": body["image"]
            }}
        ]}],
        "generationConfig": {"temperature": 0.1, "maxOutputTokens": 2048}
    }
    try:
        r = requests.post(url, json=payload,
                          headers={"Content-Type": "application/json",
                                   "x-goog-api-key": GEMINI_API_KEY}, timeout=60)
        r.raise_for_status()
        raw = r.json()["candidates"][0]["content"]["parts"][0]["text"]
        clean = raw.replace("```json", "").replace("```", "").strip()
        data = json.loads(clean)
        for it in data.get("items", []):
            it["amount"] = round(_n(it.get("qty")) * _n(it.get("price")))
        return jsonify(ok=True, data=data)
    except json.JSONDecodeError:
        return jsonify(error="Gemini 回傳非 JSON", raw=raw[:500]), 500
    except Exception as e:
        return jsonify(error=str(e)), 500


@app.route("/export/xlsx", methods=["POST"])
def export_xlsx():
    data = request.get_json(silent=True)
    if not data:
        return jsonify(error="無資料"), 400
    try:
        buf = _fill_xlsx(data)
        fname = f"報價單_{(data.get('client') or '客戶')[:10]}_{date.today()}.xlsx"
        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=fname)
    except Exception as e:
        return jsonify(error=str(e)), 500


@app.route("/export/pdf", methods=["POST"])
def export_pdf():
    data = request.get_json(silent=True)
    if not data:
        return jsonify(error="無資料"), 400
    try:
        buf = _make_pdf(data)
        fname = f"報價單_{(data.get('client') or '客戶')[:10]}_{date.today()}.pdf"
        return send_file(buf, mimetype="application/pdf",
                         as_attachment=True, download_name=fname)
    except Exception as e:
        return jsonify(error=str(e)), 500


# ═══ Excel：直接填入你的原始模板 ═══════════════════════════════════════

def _fill_xlsx(data):
    wb = load_workbook(TEMPLATE)
    ws = wb.active

    ws["B5"] = f"此致：{data.get('client', '')}"
    ws["E5"] = f"電話：{data.get('phone', '')}　　　　　傳真：{data.get('fax', '')}"
    ws["B6"] = _roc(data.get("date", ""))
    ws["E6"] = f"工地：{data.get('site', '')}"

    items = (data.get("items") or [])[:16]
    for i, it in enumerate(items):
        row = 8 + i
        ws[f"B{row}"] = it.get("name", "")
        ws[f"C{row}"] = it.get("spec", "")
        q = _n(it.get("qty"))
        p = _n(it.get("price"))
        ws[f"D{row}"] = q if q else ""
        ws[f"E{row}"] = it.get("unit", "")
        ws[f"F{row}"] = p if p else ""
        ws[f"G{row}"] = round(q * p) if q and p else ""
        ws[f"H{row}"] = it.get("note", "")

    ws["G39"] = f"報價人：{data.get('vendor', '陳\u3000世\u3000助')}"
    ws["G40"] = f"電話：{data.get('vendor_phone', '0931661180')}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══ PDF：用 ReportLab 重現同樣版面 ═══════════════════════════════════

def _make_pdf(data):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    fn = "Helvetica"
    for fp in ["/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
               "/usr/share/fonts/opentype/noto/NotoSansCJKtc-Regular.otf",
               "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
               "/usr/share/fonts/truetype/arphic/uming.ttc"]:
        if os.path.exists(fp):
            try:
                pdfmetrics.registerFont(TTFont("CJK", fp))
                fn = "CJK"; break
            except Exception:
                continue

    def S(sz, al=0, bold=False):
        return ParagraphStyle("s", fontName=fn, fontSize=sz, alignment=al,
                              leading=sz + 4, spaceAfter=1)

    items = data.get("items") or []
    sub = sum(_n(i.get("qty")) * _n(i.get("price")) for i in items)
    tax = round(sub * 0.05)
    total = round(sub + tax)
    vendor = data.get("vendor", "陳　世　助")
    vphone = data.get("vendor_phone", "0931661180")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=14*mm, rightMargin=14*mm,
                            topMargin=12*mm, bottomMargin=12*mm)
    story = []

    # 標題
    story.append(Paragraph("報　　價　　單", S(20, 1, True)))
    story.append(Paragraph("登發室內裝璜工程有限公司", S(16, 1)))
    story.append(Paragraph("台中市大雅區清陽一路50巷19號", S(11, 1)))
    story.append(Paragraph("TEL：04-25651230　　FAX：04-25651231", S(11, 1)))
    story.append(Spacer(1, 4*mm))

    # 客戶
    info = [
        [Paragraph(f"此致：{data.get('client','')}", S(11)),
         Paragraph(f"電話：{data.get('phone','')}　傳真：{data.get('fax','')}", S(11))],
        [Paragraph(_roc(data.get("date","")), S(12)),
         Paragraph(f"工地：{data.get('site','')}", S(12))]
    ]
    t0 = Table(info, colWidths=[93*mm, 87*mm])
    t0.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                            ("BOTTOMPADDING",(0,0),(-1,-1),3)]))
    story.append(t0)
    story.append(Spacer(1, 2*mm))

    # 品項表格
    cw = [8*mm, 52*mm, 22*mm, 14*mm, 11*mm, 22*mm, 22*mm, 29*mm]
    hdr = ["", "品　名", "規　格", "數量", "單位", "單　價", "金　額", "備　註"]
    tdata = [[Paragraph(h, S(10, 1)) for h in hdr]]

    for i in range(16):
        if i < len(items):
            it = items[i]
            amt = _n(it.get("qty")) * _n(it.get("price"))
            r = [str(i+1), it.get("name",""), it.get("spec",""),
                 _fmt(_n(it.get("qty"))), it.get("unit",""),
                 _fmt(_n(it.get("price"))), _fmt(amt), it.get("note","")]
        else:
            r = [str(i+1)] + [""]*7
        tdata.append([Paragraph(str(c), S(9, 1 if j!=1 else 0)) for j, c in enumerate(r)])

    tdata.append(_summary_row("小　計", _fmt(sub), S, Paragraph))
    tdata.append(_summary_row("5%營業稅", _fmt(tax), S, Paragraph))
    tdata.append(_summary_row("合　計", _fmt(total), S, Paragraph))

    tb = Table(tdata, colWidths=cw, repeatRows=1)
    tb.setStyle(TableStyle([
        ("FONTNAME",(0,0),(-1,-1), fn),
        ("FONTSIZE",(0,0),(-1,-1), 9),
        ("ALIGN",(0,0),(-1,0),"CENTER"),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("ROWHEIGHT",(0,0),(-1,-1), 6.5*mm),
        ("BACKGROUND",(0,0),(-1,0), colors.HexColor("#E0E0E0")),
        ("BOX",(0,0),(-1,-1), 1.2, colors.black),
        ("INNERGRID",(0,0),(-1,-1), 0.4, colors.black),
        ("LINEABOVE",(0,-3),(-1,-3), 1.2, colors.black),
    ]))
    story.append(tb)

    story.append(Spacer(1, 6*mm))
    st = Table([["", f"報價人：{vendor}"], ["", f"電話：{vphone}"]],
               colWidths=[120*mm, 60*mm])
    st.setStyle(TableStyle([("FONTNAME",(0,0),(-1,-1),fn),
                            ("FONTSIZE",(0,0),(-1,-1),12),
                            ("ALIGN",(1,0),(1,-1),"LEFT")]))
    story.append(st)

    doc.build(story)
    buf.seek(0)
    return buf


def _summary_row(label, value, S, Paragraph):
    return ([Paragraph("", S(9))] +
            [Paragraph(label, S(10, 1))] +
            [Paragraph("", S(9))]*4 +
            [Paragraph(value, S(10, 1))] +
            [Paragraph("", S(9))])


# ═══ 工具函式 ═════════════════════════════════════════════════════════

def _n(v):
    try: return float(v)
    except: return 0

def _fmt(n):
    n = _n(n)
    if not n: return ""
    return f"{int(n):,}" if n == int(n) else f"{n:,.1f}"

def _roc(s):
    try:
        if not s:
            t = date.today()
        else:
            p = s.split("-")
            t = date(int(p[0]), int(p[1]), int(p[2]))
        return f"報價日期：{t.year - 1911}年{t.month:02d}月{t.day:02d}日"
    except:
        return f"報價日期：{s}"


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 8080)))
